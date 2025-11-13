import math
import os
import random
from openpyxl import Workbook, load_workbook


def f(x):
    return x + (2./x) -3


def cap(a, b, eps):

    return f"f(x) = x^3 -3x + 2;\n\n[a; b] = [{a}; {b}];\n\nε = {eps}"


def random_number_between(a, b):
    while True:

        value = random.uniform(a, b)

        if value != a and value != b:
            return value


def halving_method(a, b, eps,file_path):
    if not os.path.exists(file_path):
        wb = Workbook()
        wb.save(file_path)
    else:
        wb = load_workbook(file_path)

    sheet_name = "Метод ділення навпіл"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        wb.save(file_path)

    wb.create_sheet(sheet_name)
    wb.save(file_path)

    wb[sheet_name].cell(row=1, column=1, value="K")
    wb[sheet_name].cell(row=1, column=2, value="lambda")
    wb[sheet_name].cell(row=1, column=3, value="nua")
    wb[sheet_name].cell(row=1, column=4, value="a")
    wb[sheet_name].cell(row=1, column=5, value="b")
    wb[sheet_name].cell(row=1, column=6, value="x")
    wb[sheet_name].cell(row=1, column=7, value="F(lambda)")
    wb[sheet_name].cell(row=1, column=8, value="F(nua)")
    wb[sheet_name].cell(row=1, column=9, value="F(a)")
    wb[sheet_name].cell(row=1, column=10, value="F(b)")
    wb[sheet_name].cell(row=1, column=11, value="F(x)")

    a_k = a
    b_k = b

    k = 1

    while True:

        while True:

            beta_k = random_number_between(0, b_k - a_k)

            if eps > beta_k:
                break

        lamda_k = (a_k + b_k - beta_k) / 2
        nua_k = (a_k + b_k + beta_k) / 2

        if f(lamda_k) <= f(nua_k) :
            b_k = nua_k
        else:
            a_k = lamda_k

        wb[sheet_name].cell(row=k+1, column=1, value=k)
        wb[sheet_name].cell(row=k+1, column=2, value=lamda_k)
        wb[sheet_name].cell(row=k+1, column=3, value=nua_k)
        wb[sheet_name].cell(row=k+1, column=4, value=a_k)
        wb[sheet_name].cell(row=k+1, column=5, value=b_k)
        wb[sheet_name].cell(row=k+1, column=6, value=(a_k+b_k)/2)
        wb[sheet_name].cell(row=k+1, column=7, value=f(lamda_k))
        wb[sheet_name].cell(row=k+1, column=8, value=f(nua_k))
        wb[sheet_name].cell(row=k+1, column=9, value=f(a_k))
        wb[sheet_name].cell(row=k+1, column=10, value=f(b_k))
        wb[sheet_name].cell(row=k+1, column=11, value=f((a_k+b_k)/2))


        k+=1

        if abs(b_k-a_k)<eps:
            wb.save(file_path)
            return f"f_min({(b_k + a_k) / 2}) = {f((b_k + a_k) / 2)}"


def gold_retin_method(a, b, eps,file_path):
    if not os.path.exists(file_path):
        wb = Workbook()
        wb.save(file_path)
    else:
        wb = load_workbook(file_path)

    sheet_name = "Метод Золотого перетину"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        wb.save(file_path)

    wb.create_sheet(sheet_name)
    wb.save(file_path)

    wb[sheet_name].cell(row=1, column=1, value="K")
    wb[sheet_name].cell(row=1, column=2, value="lambda")
    wb[sheet_name].cell(row=1, column=3, value="nua")
    wb[sheet_name].cell(row=1, column=4, value="a")
    wb[sheet_name].cell(row=1, column=5, value="b")
    wb[sheet_name].cell(row=1, column=6, value="x")
    wb[sheet_name].cell(row=1, column=7, value="F(lambda)")
    wb[sheet_name].cell(row=1, column=8, value="F(nua)")
    wb[sheet_name].cell(row=1, column=9, value="F(a)")
    wb[sheet_name].cell(row=1, column=10, value="F(b)")
    wb[sheet_name].cell(row=1, column=11, value="F(x)")

    a_k = a
    b_k = b

    k = 1

    while True:

        lamda_k = a_k + (3 - math.sqrt(5)) * (b_k - a_k) / 2.0
        nua_k = a_k + (math.sqrt(5) - 1) * (b_k - a_k) / 2.0

        if f(lamda_k) <= f(nua_k):
            b_k = nua_k
        else:
            a_k = lamda_k

        x_k = lamda_k if f(lamda_k) < f(nua_k) else nua_k


        wb[sheet_name].cell(row=k+1, column=1, value=k)
        wb[sheet_name].cell(row=k+1, column=2, value=lamda_k)
        wb[sheet_name].cell(row=k+1, column=3, value=nua_k)
        wb[sheet_name].cell(row=k+1, column=4, value=a_k)
        wb[sheet_name].cell(row=k+1, column=5, value=b_k)
        wb[sheet_name].cell(row=k+1, column=6, value=x_k)
        wb[sheet_name].cell(row=k+1, column=7, value=f(lamda_k))
        wb[sheet_name].cell(row=k+1, column=8, value=f(nua_k))
        wb[sheet_name].cell(row=k+1, column=9, value=f(a_k))
        wb[sheet_name].cell(row=k+1, column=10, value=f(b_k))
        wb[sheet_name].cell(row=k+1, column=11, value=f(x_k))


        k+=1

        if math.fabs(b_k-a_k)<eps:
            wb.save(file_path)
            return f"f_min({x_k}) = {f(x_k)}"


def fibonaci(x):

    if x < 3:
        return 1

    last = 1
    real = 1

    for i in range(3, x+1):

        next = last + real

        last = real

        real = next


    return real


def fibonaci_method(a, b, eps,file_path):
    if not os.path.exists(file_path):
        wb = Workbook()
        wb.save(file_path)
    else:
        wb = load_workbook(file_path)

    sheet_name = "Метод Фібоначі"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        wb.save(file_path)

    wb.create_sheet(sheet_name)
    wb.save(file_path)

    wb[sheet_name].cell(row=1, column=1, value="K")
    wb[sheet_name].cell(row=1, column=2, value="lambda")
    wb[sheet_name].cell(row=1, column=3, value="nua")
    wb[sheet_name].cell(row=1, column=4, value="a")
    wb[sheet_name].cell(row=1, column=5, value="b")
    wb[sheet_name].cell(row=1, column=6, value="x")
    wb[sheet_name].cell(row=1, column=7, value="F(lambda)")
    wb[sheet_name].cell(row=1, column=8, value="F(nua)")
    wb[sheet_name].cell(row=1, column=9, value="F(a)")
    wb[sheet_name].cell(row=1, column=10, value="F(b)")
    wb[sheet_name].cell(row=1, column=11, value="F(x)")

    a_k = a
    b_k = b

    k = 1
    n = 1

    while fibonaci(n + 1) > (b - a) / eps or (b - a) / eps > fibonaci(n + 2):
        n += 1

    while True:

        lamda_k = a_k + (b_k - a_k) * fibonaci(n - k + 1) / fibonaci(n - k + 3)
        nua_k = a_k + (b_k - a_k) * fibonaci(n - k + 2) / fibonaci(n - k + 3)

        if f(lamda_k) <= f(nua_k):
            b_k = nua_k
        else:
            a_k = lamda_k

        x_k = lamda_k if f(lamda_k) < f(nua_k) else nua_k


        wb[sheet_name].cell(row=k+1, column=1, value=k)
        wb[sheet_name].cell(row=k+1, column=2, value=lamda_k)
        wb[sheet_name].cell(row=k+1, column=3, value=nua_k)
        wb[sheet_name].cell(row=k+1, column=4, value=a_k)
        wb[sheet_name].cell(row=k+1, column=5, value=b_k)
        wb[sheet_name].cell(row=k+1, column=6, value=x_k)
        wb[sheet_name].cell(row=k+1, column=7, value=f(lamda_k))
        wb[sheet_name].cell(row=k+1, column=8, value=f(nua_k))
        wb[sheet_name].cell(row=k+1, column=9, value=f(a_k))
        wb[sheet_name].cell(row=k+1, column=10, value=f(b_k))
        wb[sheet_name].cell(row=k+1, column=11, value=f(x_k))

        if n == k:
            wb.save(file_path)
            return f"f_min({x_k}) = {f(x_k)}"

        k+=1

if __name__ == "__main__":

    a = -5
    b=-0.5
    eps=1e-4

    file_path = 'file.xlsx'

    for i in range(1000):
        print(f"\n\n\nThe Halving method: {halving_method(a, b, eps,file_path)}")
        print(f"The Gold retin method: {gold_retin_method(a, b, eps, file_path)}")
        print(f"The Fibonaci method: {fibonaci_method(a, b, eps, file_path)}")



